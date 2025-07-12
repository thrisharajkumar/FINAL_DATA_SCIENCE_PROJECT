import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

pdf_path = "C2M0160120D.pdf"
output_excel = "C2M0160120D.xlsx"

valid_headings = [
    "maximum ratings",
    "Diode Characteristics",
    "Thermal Characteristics",
    "electrical characteristics",
    "switching characteristics"
    
]
sections = []

with pdfplumber.open(pdf_path) as pdf:
    for page_num, page in enumerate(pdf.pages):
        tables = page.extract_tables()
        text_lines = page.extract_text().split("\n")

        for idx, table in enumerate(tables):
            if not table or len(table) < 2 or len(table[0]) < 2:
                continue

            found_heading = None
            for i, line in enumerate(text_lines):
                line_lower = line.strip().lower()
                if any(keyword in line_lower for keyword in valid_headings):
                    found_heading = line.strip()
                    break

            if not found_heading:
                continue  

            df = pd.DataFrame(table[1:], columns=table[0])
            sections.append((found_heading, df))
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    row_cursor = 0
    for title, df in sections:
        pd.DataFrame([[title]]).to_excel(writer, index=False, header=False, startrow=row_cursor)
        row_cursor += 1
        df.to_excel(writer, index=False, startrow=row_cursor)
        row_cursor += len(df) + 2

wb = load_workbook(output_excel)
ws = wb.active
row = 1
for title, df in sections:
    col_count = len(df.columns)
    col_letter = get_column_letter(col_count)
    ws.merge_cells(f"A{row}:{col_letter}{row}")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    ws[f"A{row}"].value = title
    row += len(df) + 3

wb.save(output_excel)
print(f"{output_excel}")
