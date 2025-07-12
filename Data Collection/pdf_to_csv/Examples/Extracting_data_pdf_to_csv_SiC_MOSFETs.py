import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Folders
input_folder = "SICMOSFETS"
output_folder = "SICMOSFET_EXCEL_FILES"
os.makedirs(output_folder, exist_ok=True)

# Keywords to match known table headers
valid_keywords = [
    "maximum ratings", "absolute maximum ratings",
    "electrical characteristics", "static characteristics",
    "switching characteristics", "switching energy", "switching times",
    "gate charge", "reverse diode", "diode characteristics", "body diode",
    "thermal characteristics", "package information", "mechanical data"
]

def find_section_heading(text_lines):
    """Return first matched known heading from text_lines (if any)."""
    for line in text_lines:
        line_clean = line.strip().lower()
        for keyword in valid_keywords:
            if keyword in line_clean:
                return line.strip()
    return None  # fallback if not found

# Iterate over all PDFs in the folder
for filename in os.listdir(input_folder):
    if not filename.lower().endswith(".pdf"):
        continue

    pdf_path = os.path.join(input_folder, filename)
    output_path = os.path.join(output_folder, filename.replace(".pdf", ".xlsx"))
    sections = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text_lines = page.extract_text().split("\n") if page.extract_text() else []
                tables = page.extract_tables()

                for idx, table in enumerate(tables):
                    if not table or len(table) < 2 or len(table[0]) < 2:
                        continue  # skip malformed tables

                    heading = find_section_heading(text_lines)
                    if not heading:
                        heading = f"Table from Page {page_num + 1} - {idx + 1}"

                    try:
                        df = pd.DataFrame(table[1:], columns=table[0])
                        sections.append((heading, df))
                    except Exception as e:
                        # Still add the raw table if columns mismatch
                        df = pd.DataFrame(table)
                        heading = f"Unstructured Table from Page {page_num + 1} - {idx + 1}"
                        sections.append((heading, df))
    except Exception as e:
        print(f" Error reading {filename}: {e}")
        continue

    # Write to Excel regardless of table quality
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        row_cursor = 0
        for title, df in sections:
            pd.DataFrame([[title]]).to_excel(writer, index=False, header=False, startrow=row_cursor)
            row_cursor += 1
            df.to_excel(writer, index=False, startrow=row_cursor)
            row_cursor += len(df) + 2

    # Merge and center titles
    wb = load_workbook(output_path)
    ws = wb.active
    row = 1
    for title, df in sections:
        col_count = len(df.columns)
        col_letter = get_column_letter(col_count)
        ws.merge_cells(f"A{row}:{col_letter}{row}")
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].value = title
        row += len(df) + 3

    wb.save(output_path)
    print(f" Processed: {filename} → {output_path}")
